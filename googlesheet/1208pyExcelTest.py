from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, Fill
from openpyxl.styles import colors
from openpyxl.styles import Fill,fills
from openpyxl.formatting.rule import ColorScaleRule

sh = load_workbook(r'C:\Users\1\Desktop\NU9_FPY_20171204-d.xlsx')
wks = sh.get_sheet_by_name('Touch_issues')
# for sheet in sh:
#     print(sheet.title)  #打印出所有sheet 的标题
# print(sh.get_sheet_names())   #打印出所有sheet 的标题
# for x in range(2, len(wks.rows)+1):
#     print(wks.cell(row=x, column=4).value) #打印出第4列所有的值
for row in wks:
    print(row[3].value)